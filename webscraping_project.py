from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import keys
from twilio.rest import Client

url = 'https://finance.yahoo.com/crypto/'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(url, headers=headers)

webpage = urlopen(req).read()

soup = BeautifulSoup(webpage,'html.parser')
print(soup.title.text)

client = Client(keys.accountSID, keys.authToken)

MyCellNum = "+16613901505"
TwilioNum = "+15074163593"

tables = soup.findAll('table')
updated_tables = tables[0]
rows = updated_tables.findAll('tr')
wb = xl.Workbook()
ws = wb.active
ws.title = "Crypto Dashboard"

HeaderFont = Font(name='Verdana',size=16,bold=True)
CellFont = Font(name='Verdana',size=14)
ws['B1'].font = HeaderFont
ws['C1'].font = HeaderFont
ws['D1'].font = HeaderFont
ws['E1'].font = HeaderFont
ws['E1'].alignment = Alignment(wrapText=True)
ws['F1'].font = HeaderFont

ws['B1'] = "Symbol"
ws['C1'] = "Currency Name"
ws['D1'] = "Current Price"
ws['E1'] = "Percent Change 24hrs"
ws['F1'] = "Price t-1"

ws.column_dimensions['B'].width = 11.5
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 20.5
ws.column_dimensions['E'].width = 34
ws.column_dimensions['E'].width = 13.33


for row in range(2,7):
    td = rows[row-1].findAll('td')
    symbol = td[0].text
    curr_name = td[1].text
    price = float(td[2].text.replace(",", "").replace("$", ""))
    percent_change = float(td[4].text.replace("%",""))
    price_tminusone = round(price - (price * percent_change / 100), 2)
    change_detector = price - price_tminusone

    if symbol in ['BTC-USD', 'ETH-USD'] and (change_detector >= 5 or change_detector <= -5):
        message = f"A price change of {change_detector} has occurred for {symbol}."
        alert = client.messages.create(to=MyCellNum, from_=TwilioNum, body=message)


    ws['B'+str(row)] = symbol
    ws['C'+str(row)] = curr_name
    ws['D'+str(row)] = '$' + str(format(price,',.2f'))
    ws['E'+str(row)] = str(format(percent_change,',.2f')+'%') 
    ws['F'+str(row)] = '$' + format(price_tminusone,',.2f')

    for col in ['B','C','D','E','F']:
        ws[col+str(row)].font = CellFont

    for col in ['B','C','D','E','F']:
        ws.column_dimensions[col].auto_size = False


wb.save('CryptoDashboard.xlsx')






