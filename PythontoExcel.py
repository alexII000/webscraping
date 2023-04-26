import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title = 'Second Sheet')

ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman', size = 24, bold=True, italic=False)

myfont = Font(name='Times New Roman', size = 24, bold=True, italic=False)

ws['A1'].font = myfont

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = myfont

ws['B8'] = '=sum(b2:b4)'

# ws.column_dimensions['A'] = 25

#Read the exvel file 'ProduceReport.xlsx' that you created earlier. Write all the contents of this file to 'Second Sheet' in the current workbook
#Display the Grand Total and Average of 'Amt Sold and 'Total' at the bottom of the list with appropriate labels

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column
maxR = read_ws.max_row

write_sheet['A1'] = 'Produce'
write_sheet['B2'] = 'Cost Per Pound'
write_sheet['C1'] = 'Amt Sold'
write_sheet['D1'] = 'Total'


wb.save('PythontoExcel.xlsx')
