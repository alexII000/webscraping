import openpyxl as xl

wb = xl.load_workbook('example.xlsx')

sn = wb.sheetnames
print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(type(cellA1.value))
print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

print(sheet1.cell(1,2).value)


print(sheet1.max_row)
print(sheet1.max_column)

maxrow = sheet1.max_row
maxcolumn = sheet1.max_column

for i in range(1, sheet1.max_row+1):
    print(sheet1.cell(i,2).value)

# n=1
# columnb = sheet1.cell(n=1,2)
# for cell in columnb:

#     print(columnb.value)

print(xl.utils.get_column_letter(1))
print(xl.utils.get_column_letter(900))

print(xl.utils.column_index_from_string('AHP'))


for currentrow in sheet1['A1':'C3']:
    print(currentrow)
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)

for currentrow in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, max_col=sheet1.max_column):
    print(currentrow)
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)

